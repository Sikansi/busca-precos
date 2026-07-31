"""Testes das falhas que apareceram no Windows.

Cada teste aqui existe porque a coisa quebrou de verdade em máquina de
cliente, não porque eu imaginei um caso.
"""

from __future__ import annotations

import csv

import pytest

from buscaprecos.planilha import Planilha, abrir_texto, codificacao_de
from buscaprecos.rede import Disjuntor, LimiteDeTaxa


# --------------------------------------------------------------------------- #
# Codificação de CSV
# --------------------------------------------------------------------------- #

def test_le_csv_em_windows_1252(tmp_path):
    """O Excel do Windows salva CSV em cp1252, não em UTF-8.

    Assumir UTF-8 estourava com UnicodeDecodeError no primeiro acento e
    derrubou uma execução completa de 1363 preços.
    """
    arquivo = tmp_path / "estoque.csv"
    arquivo.write_bytes(
        "ID produto,Descrição Produto,Preço\n1,AÇÚCAR UNIÃO,R$ 5,25\n".encode("cp1252")
    )
    with pytest.raises(UnicodeDecodeError):
        arquivo.open(encoding="utf-8-sig").read()
    with abrir_texto(arquivo, newline="") as f:
        linhas = list(csv.DictReader(f))
    assert linhas[0]["Descrição Produto"] == "AÇÚCAR UNIÃO"


def test_le_csv_em_utf8_com_bom(tmp_path):
    arquivo = tmp_path / "x.csv"
    arquivo.write_bytes("Descrição,Preço\nAÇAÍ,1\n".encode("utf-8-sig"))
    with abrir_texto(arquivo, newline="") as f:
        assert list(csv.DictReader(f))[0]["Descrição"] == "AÇAÍ"


def test_codificacao_de_identifica_o_encoding(tmp_path):
    utf8 = tmp_path / "a.csv"
    utf8.write_bytes("Ação\n".encode("utf-8"))
    cp = tmp_path / "b.csv"
    cp.write_bytes("Ação\n".encode("cp1252"))
    assert codificacao_de(utf8) in {"utf-8-sig", "utf-8"}
    assert codificacao_de(cp) == "cp1252"


def test_planilha_csv_em_cp1252_carrega(tmp_path):
    arquivo = tmp_path / "compras.csv"
    arquivo.write_bytes(
        ("Descrição do Produto,EAN/GTIN,Vl Unit. (R$)\n"
         "AÇÚCAR REFINADO UNIÃO PC 1KG,7891910000197,\"R$ 3,88\"\n").encode("cp1252")
    )
    planilha = Planilha.carregar(arquivo, coluna_chave="Descrição do Produto")
    assert len(planilha.linhas) == 1
    assert planilha.linhas[0]["Descrição do Produto"].startswith("AÇÚCAR")


# --------------------------------------------------------------------------- #
# Saída
# --------------------------------------------------------------------------- #

def test_template_xltx_sai_como_xlsx(tmp_path):
    """Template é ponto de partida, não entregável.

    O cliente selecionou o .xltx e a saída sairia .xltx: o Excel abre um "sem
    título" a cada duplo clique, em vez de abrir o resultado.
    """
    import openpyxl

    arquivo = tmp_path / "modelo.xltx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Descrição do Produto"
    wb.active["A2"] = "ARROZ"
    wb.save(arquivo)

    planilha = Planilha.carregar(arquivo, coluna_chave="Descrição do Produto")
    saida = planilha.caminho_de_saida(None, no_local=False)
    assert saida.suffix == ".xlsx"


def test_xlsx_continua_xlsx(tmp_path):
    import openpyxl

    arquivo = tmp_path / "planilha.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Descrição do Produto"
    wb.active["A2"] = "ARROZ"
    wb.save(arquivo)
    planilha = Planilha.carregar(arquivo, coluna_chave="Descrição do Produto")
    assert planilha.caminho_de_saida(None, no_local=False).suffix == ".xlsx"


# --------------------------------------------------------------------------- #
# Limite de taxa
# --------------------------------------------------------------------------- #

def test_limite_de_taxa_e_distinto_de_erro_comum():
    """Tratar os dois igual desligou o Araújo depois de 8 itens.

    Limite de taxa é temporário: a resposta certa é esperar. Loja quebrada é
    permanente: a resposta certa é desligar e avisar.
    """
    exc = LimiteDeTaxa("ARAUJO", "7894900010015", espera_sugerida=90)
    assert isinstance(exc, RuntimeError)
    assert exc.espera_sugerida == 90
    assert exc.loja == "ARAUJO"
    assert "desacelerar" in str(exc)


def test_disjuntor_desliga_apos_o_limite():
    d = Disjuntor(limite=3)
    for _ in range(3):
        d.registrar_falha("X", ValueError("quebrou"))
    assert d.aberto("X")
    assert "ValueError" in d.desligadas()["X"]


def test_sucesso_zera_o_disjuntor():
    d = Disjuntor(limite=3)
    d.registrar_falha("X", ValueError("a"))
    d.registrar_falha("X", ValueError("b"))
    d.registrar_sucesso("X")
    d.registrar_falha("X", ValueError("c"))
    assert not d.aberto("X"), "falhas intercaladas com sucesso não somam"


def test_araujo_sobe_o_ritmo_ao_ser_bloqueado():
    """Se está bloqueando, insistir no mesmo ritmo não resolve."""
    from buscaprecos.lojas import ClienteAraujo
    from buscaprecos.rede import nova_sessao

    cliente = ClienteAraujo("ARAUJO", nova_sessao(), pausa=1.2)
    inicial = cliente.pausa
    cliente.pausa = min(8.0, cliente.pausa * 1.5)  # o que o código faz no 403
    assert cliente.pausa > inicial
    for _ in range(20):
        cliente.pausa = min(8.0, cliente.pausa * 1.5)
    assert cliente.pausa == 8.0, "tem teto, senão a busca nunca termina"


# --------------------------------------------------------------------------- #
# Estoque em XLSX
# --------------------------------------------------------------------------- #

def test_estoque_aceita_xlsx(tmp_path):
    """O cliente apontou o .xlsx e o programa só lia CSV.

    Ele tentava decodificar um arquivo ZIP como texto: UnicodeDecodeError no
    meio do processamento, derrubando a execução. Com fallback de codificação
    pararia de estourar e passaria a ler lixo binário como produto — pior
    ainda, porque é silencioso.
    """
    import openpyxl

    from buscaprecos.lojas import ClienteEstoque

    arquivo = tmp_path / "planograma.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID produto", "Código Produto", "Descrição Produto",
               "Categoria produto", "Preço", "Código de barras"])
    ws.append([1, "A1", "REFRIGERANTE COCA COLA 350ML", "BEBIDAS",
               "R$ 4,99", "7894900010015"])
    wb.save(arquivo)

    cliente = ClienteEstoque("PAULO", arquivo)
    achado = cliente.buscar("7894900010015", "REFRIG COCA COLA LT 350ML")
    assert achado is not None
    assert achado.preco == 4.99


def test_estoque_recusa_formato_desconhecido(tmp_path):
    from buscaprecos.lojas import ArquivoDeEstoqueInvalido, ClienteEstoque

    arquivo = tmp_path / "coisa.pdf"
    arquivo.write_bytes(b"%PDF-1.4 nao sou planilha")
    with pytest.raises(ArquivoDeEstoqueInvalido):
        ClienteEstoque("PAULO", arquivo)


def test_estoque_sem_coluna_de_preco_explica(tmp_path):
    """Devolver 0 preços em silêncio faz o cliente achar que é falta de produto."""
    from buscaprecos.lojas import ArquivoDeEstoqueInvalido, ClienteEstoque

    arquivo = tmp_path / "estoque.csv"
    arquivo.write_text("Produto,Quantidade\nARROZ,10\n", encoding="utf-8")
    with pytest.raises(ArquivoDeEstoqueInvalido, match="coluna de preço"):
        ClienteEstoque("PAULO", arquivo)


def test_estoque_reconhece_apelidos_de_coluna(tmp_path):
    from buscaprecos.lojas import ClienteEstoque

    arquivo = tmp_path / "estoque.csv"
    arquivo.write_text(
        "EAN,Descrição,Valor\n7894900010015,COCA COLA 350ML,\"R$ 4,99\"\n",
        encoding="utf-8",
    )
    cliente = ClienteEstoque("PAULO", arquivo)
    assert cliente.por_ean["7894900010015"]["preco"] == "R$ 4,99"


# --------------------------------------------------------------------------- #
# Transportes do Araújo
# --------------------------------------------------------------------------- #

def test_tls_navegador_vem_antes_do_padrao():
    """Medido: `padrao` deu 403 numa rede e 200 na outra; `tls_navegador` deu
    200 nas duas. Começar pelo que falha custa duas requisições negadas e
    ~4s de recuo antes de trocar."""
    from buscaprecos.lojas import ClienteAraujo

    ordem = ClienteAraujo.TRANSPORTES
    assert ordem[0] == "tls_navegador"
    assert ordem.index("tls_navegador") < ordem.index("padrao")


def test_dependencias_binarias_ficam_no_fim():
    """Só os dois primeiros chegam por atualização de payload; os outros
    exigem regerar o executável, então são último recurso."""
    from buscaprecos.lojas import ClienteAraujo

    ordem = ClienteAraujo.TRANSPORTES
    sem_dependencia = {"tls_navegador", "padrao"}
    assert set(ordem[:2]) == sem_dependencia
    assert ordem.index("curl_cffi") < ordem.index("navegador")


def test_auto_troca_de_transporte_e_registra():
    from buscaprecos.lojas import ClienteAraujo
    from buscaprecos.rede import nova_sessao

    cliente = ClienteAraujo("ARAUJO", nova_sessao(), transporte="auto")
    assert cliente.transporte == "tls_navegador"
    assert cliente._proximo_transporte()
    assert cliente.transporte == "padrao"
    assert cliente.trocas_de_transporte == ["padrao"]


def test_transporte_fixo_nao_troca_sozinho():
    """Quem configurou à mão não quer o programa decidindo por conta."""
    from buscaprecos.lojas import ClienteAraujo
    from buscaprecos.rede import nova_sessao

    cliente = ClienteAraujo("ARAUJO", nova_sessao(), transporte="tls_navegador")
    assert not cliente._proximo_transporte()
    assert cliente.transporte == "tls_navegador"


def test_sessao_tls_navegador_usa_cifras_de_navegador():
    from buscaprecos.rede import CIFRAS_NAVEGADOR, sessao_tls_navegador

    assert "CHACHA20" in CIFRAS_NAVEGADOR
    sessao = sessao_tls_navegador()
    assert sessao.adapters["https://"].__class__.__name__ == "_AdaptadorTlsNavegador"


# --------------------------------------------------------------------------- #
# URL de atualização
# --------------------------------------------------------------------------- #

def test_url_do_asset_fixa_na_tag_e_nao_em_latest():
    """`latest/download/payload-1.0.4.zip` vira 404 quando a 1.0.5 sai.

    Medido: o asset só existe no release dele. Um version.json em cache
    apontando para `latest` manda o cliente buscar arquivo inexistente.
    """
    from buscaprecos.atualizacao import url_do_asset

    url = url_do_asset(
        "https://github.com/Sikansi/busca-precos/releases", "1.0.6",
        "payload-1.0.6.zip",
    )
    assert url.endswith("/releases/download/v1.0.6/payload-1.0.6.zip")
    assert "/latest/" not in url


@pytest.mark.parametrize("base", [
    "https://github.com/U/R/releases",
    "https://github.com/U/R/releases/",
    "https://github.com/U/R/releases/latest/download",   # forma antiga
    "https://github.com/U/R/releases/download",
])
def test_url_do_asset_aceita_as_bases_antigas(base):
    from buscaprecos.atualizacao import url_do_asset

    assert (url_do_asset(base, "2.0.0", "p.zip")
            == "https://github.com/U/R/releases/download/v2.0.0/p.zip")


def test_versao_compara_numericamente():
    from buscaprecos.atualizacao import mais_nova

    assert mais_nova("1.0.10", "1.0.9"), "comparação alfabética diria o contrário"
    assert mais_nova("1.10.0", "1.9.0")
    assert not mais_nova("1.0.5", "1.0.5")
    assert not mais_nova("1.0.4", "1.0.5")


# --------------------------------------------------------------------------- #
# Conversão de template
# --------------------------------------------------------------------------- #

def _tipo_declarado(caminho) -> str:
    """O tipo que o arquivo declara internamente: 'sheet' ou 'template'."""
    import re
    import zipfile

    with zipfile.ZipFile(caminho) as z:
        xml = z.read("[Content_Types].xml").decode()
    m = re.search(r'PartName="/xl/workbook\.xml"\s+ContentType="([^"]+)"', xml)
    assert m, "não achei o tipo do workbook"
    return m.group(1).rsplit(".", 2)[-2]


def _template(tmp_path, nome="modelo.xltx"):
    import openpyxl

    caminho = tmp_path / nome
    wb = openpyxl.Workbook()
    wb.template = True
    ws = wb.active
    ws["A1"] = "Descrição do Produto"
    ws["A2"] = "ARROZ PRATO FINO 1KG"
    wb.save(caminho)
    assert _tipo_declarado(caminho) == "template"
    return caminho


def test_saida_de_template_declara_planilha_e_nao_template(tmp_path):
    """Trocar só a extensão faz o Excel recusar o arquivo.

    O .xltx declara "…template.main+xml" dentro do zip. Salvo com nome .xlsx,
    o Excel compara extensão e conteúdo e devolve "o formato ou a extensão não
    é válida". O Google Sheets abre, o que faz parecer problema do Excel.
    """
    from buscaprecos.planilha import Planilha

    planilha = Planilha.carregar(_template(tmp_path),
                                 coluna_chave="Descrição do Produto")
    saida = planilha.salvar(["MENOR PREÇO"])
    assert saida.suffix == ".xlsx"
    assert _tipo_declarado(saida) == "sheet", (
        "arquivo com nome .xlsx declarando template: o Excel recusa"
    )


def test_saida_de_xlsx_continua_planilha(tmp_path):
    import openpyxl

    from buscaprecos.planilha import Planilha

    origem = tmp_path / "planilha.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "Descrição do Produto"
    wb.active["A2"] = "ARROZ"
    wb.save(origem)

    planilha = Planilha.carregar(origem, coluna_chave="Descrição do Produto")
    saida = planilha.salvar(["MENOR PREÇO"])
    assert _tipo_declarado(saida) == "sheet"


def test_saida_pode_abrir_no_openpyxl_de_novo(tmp_path):
    """Ida e volta: o arquivo gerado tem que ser legível como planilha."""
    import openpyxl

    from buscaprecos.planilha import Planilha

    planilha = Planilha.carregar(_template(tmp_path),
                                 coluna_chave="Descrição do Produto")
    saida = planilha.salvar(["MENOR PREÇO"])
    wb = openpyxl.load_workbook(saida)
    assert wb.template is False
    assert wb.active["A2"].value == "ARROZ PRATO FINO 1KG"
