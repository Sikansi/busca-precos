"""Leitura e escrita da planilha, em XLSX ou CSV.

Regras que existem para não destruir o trabalho do cliente:

1. **Nunca sobrescreve a entrada** por padrão — grava em arquivo novo com
   data/hora. `--no-local` só com `backup` automático.
2. **Não pisa em fórmula.** Se a célula de destino já tem `=...`, ela é do
   cliente e fica como está (o caso é registrado em `avisos`). O template
   `.xltx` tem 857 fórmulas em `Tabela_1`; o ida-e-volta por CSV as achatava
   em valor estático.
3. **Escreve no tipo da coluna.** `MARKUP ALVO` e `MARGEM` são números na
   planilha (0,6 e 0,98 formatados como porcentagem), não texto "60%".
   Gravar string ali quebraria a formatação e qualquer fórmula que dependa
   delas.
4. **Só escreve nas colunas que são do programa.** Nenhuma outra é tocada.
"""

from __future__ import annotations

import csv
import re
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

MAX_LINHAS_HEADER = 10


def _agora() -> str:
    return datetime.now().strftime("%Y%m%d-%H%M")


def _texto(valor: Any) -> str:
    """Valor de célula como texto, sem inventar decimal.

    O Excel guarda tudo numérico como float: EAN vira 7891910000197.0 e
    código interno vira 60413.0. Deixar o ".0" quebra qualquer busca por
    código de barras.
    """
    if valor is None:
        return ""
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


# O Excel do Windows salva CSV em Windows-1252 por padrão, não em UTF-8.
# Assumir UTF-8 estoura com UnicodeDecodeError no primeiro acento — foi o que
# derrubou uma execução completa de 1363 preços. latin-1 vai por último porque
# aceita qualquer byte: nunca falha, no pior caso mostra acento errado, o que é
# infinitamente melhor que perder o trabalho.
CODIFICACOES = ("utf-8-sig", "utf-8", "cp1252", "latin-1")


def abrir_texto(caminho: Path, **kwargs: Any) -> Any:
    """Abre arquivo de texto tentando as codificações usuais em ordem."""
    ultimo_erro: UnicodeDecodeError | None = None
    for codificacao in CODIFICACOES:
        try:
            f = caminho.open(encoding=codificacao, **kwargs)
            f.read(8192)
            f.seek(0)
            return f
        except UnicodeDecodeError as exc:
            ultimo_erro = exc
            try:
                f.close()
            except Exception:
                pass
    raise ultimo_erro or UnicodeDecodeError(
        "utf-8", b"", 0, 1, f"não consegui ler {caminho.name}"
    )


def codificacao_de(caminho: Path) -> str:
    """Qual codificação funciona para este arquivo (para relatar ao usuário)."""
    for codificacao in CODIFICACOES:
        try:
            with caminho.open(encoding=codificacao) as f:
                f.read()
            return codificacao
        except UnicodeDecodeError:
            continue
    return "desconhecida"


def _eh_formula(valor: Any) -> bool:
    return isinstance(valor, str) and valor.startswith("=")


def _numero_de_texto(valor: str) -> float | None:
    """"R$ 1.234,56" → 1234.56 | "98%" → 0.98 | "12" → 12.0."""
    v = valor.strip()
    if not v:
        return None
    if v.endswith("%"):
        m = re.search(r"-?\d+(?:[.,]\d+)?", v)
        return float(m.group(0).replace(",", ".")) / 100 if m else None
    m = re.search(r"-?\d{1,3}(?:\.\d{3})*,\d{2}|-?\d+,\d{2}|-?\d+(?:\.\d+)?", v)
    if not m:
        return None
    bruto = m.group(0)
    if "," in bruto:
        return float(bruto.replace(".", "").replace(",", "."))
    return float(bruto)


@dataclass
class Planilha:
    caminho: Path
    formato: str  # "xlsx" | "csv"
    colunas: list[str]
    linhas: list[dict[str, Any]]
    avisos: list[str] = field(default_factory=list)
    _wb: Any = None
    _ws: Any = None
    _linha_header: int = 1
    _colunas_numericas: set[str] = field(default_factory=set)

    # ------------------------------------------------------------------ #
    # Carga
    # ------------------------------------------------------------------ #

    @classmethod
    def carregar(cls, caminho: Path | str, *, coluna_chave: str) -> Planilha:
        caminho = Path(caminho)
        if caminho.suffix.lower() in {".xlsx", ".xlsm", ".xltx"}:
            return cls._carregar_xlsx(caminho, coluna_chave)
        return cls._carregar_csv(caminho, coluna_chave)

    @classmethod
    def _carregar_xlsx(cls, caminho: Path, coluna_chave: str) -> Planilha:
        import openpyxl

        wb = openpyxl.load_workbook(caminho)  # com fórmulas
        wb_val = openpyxl.load_workbook(caminho, data_only=True)  # valores em cache
        ws = wb[wb.sheetnames[0]]
        ws_val = wb_val[wb_val.sheetnames[0]]

        linha_header = cls._achar_header(ws, coluna_chave)
        colunas = [
            _texto(c.value) for c in ws[linha_header]
        ]
        # remove colunas sem nome no fim
        while colunas and not colunas[-1]:
            colunas.pop()

        linhas: list[dict[str, Any]] = []
        avisos: list[str] = []
        for r in range(linha_header + 1, ws.max_row + 1):
            registro: dict[str, Any] = {}
            vazia = True
            for idx, nome in enumerate(colunas, start=1):
                if not nome:
                    continue
                bruto = ws.cell(row=r, column=idx).value
                # Para cálculo usa o valor em cache; a fórmula fica no arquivo.
                valor = ws_val.cell(row=r, column=idx).value if _eh_formula(bruto) else bruto
                registro[nome] = _texto(valor)
                if registro[nome]:
                    vazia = False
            if vazia:
                continue
            registro["__linha_excel__"] = r
            linhas.append(registro)

        numericas = cls._colunas_numericas_de(ws_val, colunas, linha_header)
        if ws_val.max_row > 1 and all(
            ws_val.cell(row=2, column=i).value is None for i in range(1, len(colunas) + 1)
        ):
            avisos.append(
                "planilha sem valores em cache — abra e salve no Excel uma vez "
                "para que as fórmulas tenham valor calculado"
            )

        return cls(
            caminho=caminho,
            formato="xlsx",
            colunas=[c for c in colunas if c],
            linhas=linhas,
            avisos=avisos,
            _wb=wb,
            _ws=ws,
            _linha_header=linha_header,
            _colunas_numericas=numericas,
        )

    @staticmethod
    def _achar_header(ws: Any, coluna_chave: str) -> int:
        """Header é a primeira linha que contém a coluna-chave."""
        alvo = coluna_chave.strip().lower()
        for r in range(1, min(MAX_LINHAS_HEADER, ws.max_row) + 1):
            for c in ws[r]:
                if _texto(c.value).lower() == alvo:
                    return r
        return 1

    @staticmethod
    def _colunas_numericas_de(
        ws_val: Any, colunas: list[str], linha_header: int
    ) -> set[str]:
        """Colunas cujo conteúdo atual é numérico — devem receber número."""
        numericas: set[str] = set()
        limite = min(ws_val.max_row, linha_header + 30)
        for idx, nome in enumerate(colunas, start=1):
            if not nome:
                continue
            num = txt = 0
            for r in range(linha_header + 1, limite + 1):
                v = ws_val.cell(row=r, column=idx).value
                if v is None or v == "":
                    continue
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    num += 1
                else:
                    txt += 1
            if num and num >= txt:
                numericas.add(nome)
        return numericas

    @classmethod
    def _carregar_csv(cls, caminho: Path, coluna_chave: str) -> Planilha:
        """Aceita header na 1ª linha (formato atual) ou na 4ª (legado)."""
        with abrir_texto(caminho, newline="") as f:
            amostra = f.read(8192)
            f.seek(0)
            primeira = amostra.splitlines()[0] if amostra else ""
            if coluna_chave in primeira:
                leitor = csv.DictReader(f)
                colunas = [c for c in (leitor.fieldnames or []) if c]
                linhas = [
                    {k: _texto(v) for k, v in row.items() if k is not None}
                    for row in leitor
                ]
            else:
                cru = list(csv.reader(f))
                colunas = [h.strip() for h in cru[3]] if len(cru) > 3 else []
                linhas = []
                for raw in cru[4:]:
                    if not raw or all(not _texto(c) for c in raw):
                        continue
                    linhas.append({
                        colunas[i]: _texto(raw[i])
                        for i in range(min(len(colunas), len(raw)))
                        if colunas[i]
                    })
        return cls(
            caminho=caminho,
            formato="csv",
            colunas=[c for c in colunas if c],
            linhas=linhas,
        )

    # ------------------------------------------------------------------ #
    # Estrutura
    # ------------------------------------------------------------------ #

    def garantir_colunas(self, nomes: list[str], *, depois_de: list[str] | None = None) -> None:
        """Cria colunas que faltam, tentando posicioná-las após uma âncora."""
        posicao: int | None = None
        for ancora in depois_de or []:
            if ancora in self.colunas:
                posicao = self.colunas.index(ancora) + 1
                break
        for nome in nomes:
            if nome in self.colunas:
                continue
            if posicao is None:
                self.colunas.append(nome)
            else:
                self.colunas.insert(posicao, nome)
                posicao += 1
        for linha in self.linhas:
            for nome in self.colunas:
                linha.setdefault(nome, "")

    def limpar_erros_excel(self, coluna: str) -> None:
        """`#DIV/0!` e afins viram vazio para não poluir o cálculo."""
        ruins = {"#DIV/0!", "#N/A", "#REF!", "#VALUE!", "#NAME?", "#NULL!"}
        for linha in self.linhas:
            if _texto(linha.get(coluna)) in ruins:
                linha[coluna] = ""

    # ------------------------------------------------------------------ #
    # Escrita
    # ------------------------------------------------------------------ #

    def caminho_de_saida(self, destino: Path | str | None, *, no_local: bool) -> Path:
        if destino:
            return Path(destino)
        if no_local:
            return self.caminho
        stem = self.caminho.stem
        sufixo = self.caminho.suffix
        # Template (.xltx/.xltm) é ponto de partida, não entregável: o Excel
        # abre um "sem título" a cada duplo clique. A saída sai como planilha.
        if sufixo.lower() in {".xltx", ".xltm"}:
            sufixo = ".xlsx"
        return self.caminho.with_name(f"{stem}_precos_{_agora()}{sufixo}")

    def salvar(
        self,
        colunas_do_programa: list[str],
        *,
        destino: Path | str | None = None,
        no_local: bool = False,
    ) -> Path:
        saida = self.caminho_de_saida(destino, no_local=no_local)
        if saida == self.caminho:
            reserva = self.caminho.with_name(
                f"{self.caminho.stem}.bak-{_agora()}{self.caminho.suffix}"
            )
            shutil.copy2(self.caminho, reserva)
            self.avisos.append(f"cópia de segurança: {reserva.name}")
        if self.formato == "xlsx":
            self._salvar_xlsx(saida, colunas_do_programa)
        else:
            self._salvar_csv(saida)
        return saida

    def _salvar_xlsx(self, saida: Path, colunas_do_programa: list[str]) -> None:
        indices = {
            _texto(c.value): c.column
            for c in self._ws[self._linha_header]
            if _texto(c.value)
        }
        # Cria no cabeçalho as colunas novas que ainda não existem na aba.
        proxima = max(indices.values(), default=0) + 1
        for nome in colunas_do_programa:
            if nome not in indices:
                self._ws.cell(row=self._linha_header, column=proxima, value=nome)
                indices[nome] = proxima
                proxima += 1

        formulas_preservadas = 0
        for linha in self.linhas:
            r = linha.get("__linha_excel__")
            if not r:
                continue
            for nome in colunas_do_programa:
                idx = indices.get(nome)
                if idx is None:
                    continue
                celula = self._ws.cell(row=r, column=idx)
                if _eh_formula(celula.value):
                    formulas_preservadas += 1
                    continue
                celula.value = self._valor_para_celula(nome, _texto(linha.get(nome)))

        if formulas_preservadas:
            self.avisos.append(
                f"{formulas_preservadas} célula(s) com fórmula preservada(s) — "
                "o cálculo do Excel tem prioridade"
            )
        self._wb.save(saida)

    def _valor_para_celula(self, coluna: str, texto: str) -> Any:
        """Converte para número quando a coluna é numérica na planilha."""
        if texto == "":
            return None
        if coluna in self._colunas_numericas:
            num = _numero_de_texto(texto)
            if num is not None:
                return num
        return texto

    def _salvar_csv(self, saida: Path) -> None:
        with saida.open("w", newline="", encoding="utf-8-sig") as f:
            escritor = csv.DictWriter(f, fieldnames=self.colunas, extrasaction="ignore")
            escritor.writeheader()
            escritor.writerows(self.linhas)


# --------------------------------------------------------------------------- #
# Detecção de colunas
# --------------------------------------------------------------------------- #

# Nomes que o cliente pode ter usado para a mesma coisa. A planilha não precisa
# bater exatamente: só `descricao` é indispensável, e mesmo ela pode ser
# apontada à mão na interface se o nome for inesperado.
APELIDOS: dict[str, tuple[str, ...]] = {
    "descricao": (
        "descricao do produto", "descricao produto", "descricao", "produto",
        "produtos", "nome do produto", "item", "mercadoria",
    ),
    "ean": (
        "ean gtin", "ean", "gtin", "codigo de barras", "cod barras",
        "cod de barras", "codigo barras", "ean13", "gtin ean", "barras",
    ),
    "codigo_interno": (
        "codigo interno", "codigo produto", "cod interno", "cod produto",
        "sku", "id produto", "codigo",
    ),
    "custo": (
        "vl unit r", "vl unit", "valor unitario", "vlr unitario", "custo",
        "custo unitario", "preco de custo", "preco custo", "vl unitario",
    ),
    "markup_planilha": ("markup", "markup alvo", "marcacao", "margem alvo"),
}

# Sem estas o programa até roda, mas perde função — a interface avisa o quê.
CONSEQUENCIA_SE_FALTAR = {
    "ean": "sem código de barras a busca cai para texto, com mais erro",
    "codigo_interno": "sem código interno o preço da própria loja acha menos",
    "custo": "sem custo não dá para calcular MARGEM, PREÇO PERTIN nem REGRA",
    "markup_planilha": "sem markup na planilha ele é deduzido da categoria",
}


def _chave(nome: str) -> str:
    """Normaliza cabeçalho para comparar: sem acento, sem pontuação, minúsculo."""
    import unicodedata

    t = unicodedata.normalize("NFKD", str(nome or ""))
    t = "".join(c for c in t if not unicodedata.combining(c)).lower()
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", t)).strip()


def detectar_colunas(
    colunas: list[str], atual: dict[str, str] | None = None
) -> tuple[dict[str, str], list[str]]:
    """Casa os papéis com os cabeçalhos da planilha.

    Devolve (mapa papel→cabeçalho, papéis que não foram encontrados).
    Tenta, em ordem: o nome já configurado, igualdade normalizada, e por
    último cabeçalho que comece com o apelido (pega "Vl Unit. (R$)" a partir
    de "vl unit").
    """
    presentes = {_chave(c): c for c in colunas if str(c).strip()}
    mapa: dict[str, str] = {}
    faltando: list[str] = []

    for papel, apelidos in APELIDOS.items():
        configurado = (atual or {}).get(papel)
        if configurado and configurado in colunas:
            mapa[papel] = configurado
            continue

        achado = None
        for apelido in apelidos:
            if apelido in presentes:
                achado = presentes[apelido]
                break
        if achado is None:
            for apelido in apelidos:
                for chave, original in presentes.items():
                    if chave.startswith(apelido):
                        achado = original
                        break
                if achado:
                    break
        if achado:
            mapa[papel] = achado
        else:
            faltando.append(papel)
    return mapa, faltando


def cabecalhos_de(caminho: Path | str) -> list[str]:
    """Só os cabeçalhos, sem carregar a planilha inteira.

    Serve para a interface montar a tela de mapeamento rápido, antes de
    decidir se dá para prosseguir.
    """
    caminho = Path(caminho)
    if caminho.suffix.lower() in {".xlsx", ".xlsm", ".xltx"}:
        import openpyxl

        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        melhor: list[str] = []
        for linha in ws.iter_rows(min_row=1, max_row=MAX_LINHAS_HEADER, values_only=True):
            nomes = [_texto(v) for v in (linha or ())]
            if sum(1 for n in nomes if n) > sum(1 for n in melhor if n):
                melhor = nomes
        wb.close()
        return [n for n in melhor if n]
    with abrir_texto(caminho, newline="") as f:
        for linha in csv.reader(f):
            if sum(1 for c in linha if str(c).strip()) >= 3:
                return [str(c).strip() for c in linha if str(c).strip()]
    return []
